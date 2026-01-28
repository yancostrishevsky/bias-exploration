"""Journal Impact Factor lookup helpers."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

from openpyxl import load_workbook

from ai_bias_search.utils.config import ImpactFactorConfig
from ai_bias_search.utils.logging import configure_logging

LOGGER = configure_logging()

PUNCT_RE = re.compile(r"[\\.,:;()\\[\\]{}'\"/\\\\-]")
SPACE_RE = re.compile(r"\s+")
HEADER_SPACE_RE = re.compile(r"\s+")
ISSN_CLEAN_RE = re.compile(r"[^0-9X]")
JIF_YEAR_HEADER_RE = re.compile(r"^jif\s+(\d{4})$")

JCR_PAYLOAD_FIELDS = (
    "impact_factor",
    "impact_factor_year",
    "jcr_year",
    "jcr_publisher",
    "jcr_issn",
    "jcr_eissn",
    "jcr_total_cites",
    "jcr_total_articles",
    "jcr_citable_items",
    "jcr_jif_5y",
    "jcr_jif_wo_self_cites",
    "jcr_jci",
    "jcr_quartile",
    "jcr_jif_rank",
)

try:  # optional fuzzy matching
    from rapidfuzz import fuzz, process  # type: ignore

    RAPIDFUZZ_AVAILABLE = True
except ImportError:  # pragma: no cover - optional dependency
    RAPIDFUZZ_AVAILABLE = False
    fuzz = None  # type: ignore[assignment]
    process = None  # type: ignore[assignment]


@dataclass(frozen=True)
class ImpactFactorIndex:
    """Lookup tables for journal impact factors."""

    by_title: Dict[str, Dict[str, object]]
    by_issn: Dict[str, Dict[str, object]]
    title_keys: Tuple[str, ...]
    issn_keys: Tuple[str, ...]


def normalize_journal_title(title: str) -> str | None:
    """Normalize a journal title for matching."""

    if not title or not isinstance(title, str):
        return None
    text = title.strip().lower()
    if not text:
        return None
    text = text.replace("&", " and ")
    text = PUNCT_RE.sub(" ", text)
    text = SPACE_RE.sub(" ", text)
    return text.strip() or None


def normalize_issn(value: str | None) -> str | None:
    """Normalize ISSN/EISSN to a stable key."""

    if value is None:
        return None
    raw_text = str(value).strip().upper()
    if not raw_text:
        return None
    text = ISSN_CLEAN_RE.sub("", raw_text)
    if len(text) == 8:
        return f"{text[:4]}-{text[4:]}"
    if len(text) == 7 and re.fullmatch(r"[0-9X]{7}", raw_text):
        text = f"0{text}"
        return f"{text[:4]}-{text[4:]}"
    return None


def load_jif_xlsx(
    path: Path,
    *,
    sheet_name: str | None,
    title_column: str,
    jif_column: str,
    year_column: str | None = None,
    publisher_column: str | None = None,
    issn_column: str | None = None,
    eissn_column: str | None = None,
    total_cites_column: str | None = None,
    citable_items_column: str | None = None,
    total_articles_column: str | None = None,
    jif_5y_column: str | None = None,
    jif_wo_self_cites_column: str | None = None,
    jci_column: str | None = None,
    quartile_column: str | None = None,
    jif_rank_column: str | None = None,
) -> ImpactFactorIndex:
    """Load impact factors from an XLSX file."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            if sheet_name:
                LOGGER.warning("JIF sheet %r not found in %s; using active sheet", sheet_name, path)
            sheet = workbook.active

        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            LOGGER.warning("JIF workbook %s has no header row", path)
            return ImpactFactorIndex({}, {}, (), ())

        headers = [str(cell).strip() if cell is not None else "" for cell in header]
        title_idx, title_name_used = _resolve_required_column(headers, title_column, kind="title")
        jif_idx, jif_name_used = _resolve_required_column(headers, jif_column, kind="jif")

        if title_idx is None or jif_idx is None:
            raise ValueError(
                f"Missing required JIF columns: {title_column!r} (resolved={title_name_used!r}), "
                f"{jif_column!r} (resolved={jif_name_used!r})"
            )

        year_idx: int | None = None
        if year_column:
            year_idx, year_name_used = _resolve_optional_column(headers, year_column, kind="year")
            if year_idx is None:
                LOGGER.warning(
                    "JIF year column %r not found in %s; year values omitted", year_column, path
                )
            elif year_name_used != year_column:
                LOGGER.warning(
                    "JIF year column %r not found in %s; using %r",
                    year_column,
                    path,
                    year_name_used,
                )

        optional_columns = {
            "publisher": publisher_column,
            "issn": issn_column,
            "eissn": eissn_column,
            "total_cites": total_cites_column,
            "citable_items": citable_items_column,
            "total_articles": total_articles_column,
            "jif_5y": jif_5y_column,
            "jif_wo_self_cites": jif_wo_self_cites_column,
            "jci": jci_column,
            "quartile": quartile_column,
            "jif_rank": jif_rank_column,
        }
        optional_indices: Dict[str, Optional[int]] = {}
        for key, column in optional_columns.items():
            idx = _find_column(headers, column) if column else None
            if column and idx is None:
                LOGGER.warning(
                    "JIF column %r not found in %s; %s values omitted", column, path, key
                )
            optional_indices[key] = idx

        by_title: Dict[str, Dict[str, object]] = {}
        by_issn: Dict[str, Dict[str, object]] = {}
        ambiguous_titles: set[str] = set()
        ambiguous_issn: set[str] = set()
        total_rows = 0
        loaded_titles = 0
        loaded_issn = 0

        for row in rows:
            if not row:
                continue
            total_rows += 1
            raw_title = _safe_cell(row, title_idx)
            title_text = raw_title.strip() if isinstance(raw_title, str) else ""
            if not title_text:
                continue
            title_key = normalize_journal_title(title_text)
            if not title_key:
                continue

            jif_value = _coerce_float(_safe_cell(row, jif_idx))
            if jif_value is None:
                continue

            year_value = None
            if year_idx is not None:
                year_value = _coerce_int(_safe_cell(row, year_idx))

            publisher = _coerce_str(_safe_cell(row, optional_indices["publisher"]))
            issn_raw = _coerce_str(_safe_cell(row, optional_indices["issn"]))
            eissn_raw = _coerce_str(_safe_cell(row, optional_indices["eissn"]))
            issn_value = normalize_issn(issn_raw)
            eissn_value = normalize_issn(eissn_raw)
            total_cites = _coerce_int(_safe_cell(row, optional_indices["total_cites"]))
            total_articles = _coerce_int(_safe_cell(row, optional_indices["total_articles"]))
            citable_items = _coerce_int(_safe_cell(row, optional_indices["citable_items"]))
            jif_5y = _coerce_float(_safe_cell(row, optional_indices["jif_5y"]))
            jif_wo_self_cites = _coerce_float(
                _safe_cell(row, optional_indices["jif_wo_self_cites"])
            )
            jci = _coerce_float(_safe_cell(row, optional_indices["jci"]))
            quartile = _coerce_str(_safe_cell(row, optional_indices["quartile"]))
            jif_rank = _coerce_str(_safe_cell(row, optional_indices["jif_rank"]))

            payload = {
                "raw_title": title_text,
                "impact_factor": jif_value,
                "impact_factor_year": year_value,
                "jcr_year": year_value,
                "jcr_publisher": publisher,
                "jcr_issn": issn_value,
                "jcr_eissn": eissn_value,
                "jcr_total_cites": total_cites,
                "jcr_total_articles": total_articles,
                "jcr_citable_items": citable_items,
                "jcr_jif_5y": jif_5y,
                "jcr_jif_wo_self_cites": jif_wo_self_cites,
                "jcr_jci": jci,
                "jcr_quartile": quartile,
                "jcr_jif_rank": jif_rank,
            }

            if _update_index(by_title, title_key, payload, ambiguous_titles, "title"):
                loaded_titles += 1
            if issn_value and _update_index(by_issn, issn_value, payload, ambiguous_issn, "issn"):
                loaded_issn += 1
            if eissn_value and _update_index(
                by_issn, eissn_value, payload, ambiguous_issn, "eissn"
            ):
                loaded_issn += 1

        index = ImpactFactorIndex(
            by_title=by_title,
            by_issn=by_issn,
            title_keys=tuple(by_title.keys()),
            issn_keys=tuple(by_issn.keys()),
        )
        LOGGER.info(
            "Loaded JIF index from %s (rows=%d, titles=%d, issn=%d)",
            path,
            total_rows,
            loaded_titles,
            loaded_issn,
        )
        return index
    finally:
        workbook.close()


def match_jcr_entry(
    title: str | None,
    issn_candidates: Iterable[str | None],
    index: ImpactFactorIndex,
    config: ImpactFactorConfig,
) -> Dict[str, object]:
    """Match a journal title/ISSN to a JCR entry."""

    title_key = normalize_journal_title(title) if title else None
    issn_keys = _normalize_issn_candidates(issn_candidates)

    for issn_key in issn_keys:
        match = index.by_issn.get(issn_key)
        if match:
            return _finalize_payload(match, title, title_key, "issn_exact")

    if not title_key:
        return _finalize_payload(None, title, title_key, "none")
    if len(title_key) < config.min_title_len:
        return _finalize_payload(None, title, title_key, "none")

    exact = index.by_title.get(title_key)
    if exact:
        return _finalize_payload(exact, title, title_key, "title_exact")

    if not config.allow_fuzzy:
        return _finalize_payload(None, title, title_key, "none")
    if not RAPIDFUZZ_AVAILABLE:
        LOGGER.warning("RapidFuzz not available; skipping fuzzy JIF match")
        return _finalize_payload(None, title, title_key, "none")
    if not index.title_keys:
        return _finalize_payload(None, title, title_key, "none")

    matches = process.extract(title_key, index.title_keys, scorer=fuzz.ratio, limit=2)
    if not matches:
        return _finalize_payload(None, title, title_key, "none")

    best_key, best_score, _ = matches[0]
    best_score = float(best_score)
    if best_score < config.fuzzy_threshold:
        return _finalize_payload(None, title, title_key, "none")

    best_key_str = str(best_key)
    if not _length_ratio_ok(title_key, best_key_str, config.max_len_ratio_delta):
        return _finalize_payload(None, title, title_key, "none")

    if config.reject_ambiguous and len(matches) > 1:
        second_score = float(matches[1][1])
        if best_score - second_score < 2:
            return _finalize_payload(None, title, title_key, "ambiguous")

    match = index.by_title.get(best_key_str)
    if not match:
        return _finalize_payload(None, title, title_key, "none")
    return _finalize_payload(match, title, title_key, "title_fuzzy")


def match_impact_factor(
    title: str | None,
    index: ImpactFactorIndex,
    config: ImpactFactorConfig,
) -> tuple[float | None, int | None, str, str | None]:
    """Match a journal title to a JIF entry."""

    payload = match_jcr_entry(title, [], index, config)
    return (
        payload.get("impact_factor"),  # type: ignore[return-value]
        payload.get("impact_factor_year"),  # type: ignore[return-value]
        str(payload.get("impact_factor_match") or "none"),
        payload.get("impact_factor_title_key"),  # type: ignore[return-value]
    )


def _finalize_payload(
    match: Dict[str, object] | None,
    title: str | None,
    title_key: str | None,
    match_type: str,
) -> Dict[str, object]:
    payload = _empty_payload()
    if match:
        payload.update(_payload_for_record(match))
    payload.update(
        {
            "impact_factor_match": match_type,
            "jcr_match_type": match_type,
            "impact_factor_title_raw": title,
            "impact_factor_title_key": title_key,
            "impact_factor_source": "xlsx_jif",
        }
    )
    return payload


def _empty_payload() -> Dict[str, object]:
    return {key: None for key in JCR_PAYLOAD_FIELDS}


def _payload_for_record(payload: Dict[str, object]) -> Dict[str, object]:
    return {key: payload.get(key) for key in JCR_PAYLOAD_FIELDS}


def _find_column(headers: list[str], name: str | None) -> Optional[int]:
    if not name:
        return None
    target = _normalize_header(name)
    if not target:
        return None
    for idx, header in enumerate(headers):
        if _normalize_header(header) == target:
            return idx
    return None


def _resolve_required_column(
    headers: list[str], name: str, *, kind: str
) -> tuple[int | None, str | None]:
    idx, used = _resolve_optional_column(headers, name, kind=kind)
    if idx is not None:
        return idx, used
    return None, used


def _resolve_optional_column(
    headers: list[str], name: str, *, kind: str
) -> tuple[int | None, str | None]:
    idx = _find_column(headers, name)
    if idx is not None:
        return idx, name

    normalized = _normalize_header(name)

    if kind == "title":
        for alias in ("Journal Name", "Journal Title", "Source Title"):
            alias_idx = _find_column(headers, alias)
            if alias_idx is not None:
                LOGGER.warning("JIF title column %r not found; using %r", name, alias)
                return alias_idx, alias

    if kind == "year":
        for alias in ("JCR Year", "Year"):
            if _normalize_header(alias) == normalized:
                continue
            alias_idx = _find_column(headers, alias)
            if alias_idx is not None:
                return alias_idx, alias

    if kind == "jif":
        # Common JCR export uses "JIF <year>" (e.g., "JIF 2024"). If the config asks for a generic
        # "JIF", pick the newest available year column.
        if normalized == "jif":
            candidates: list[tuple[int, int]] = []
            for idx2, header in enumerate(headers):
                match = JIF_YEAR_HEADER_RE.match(_normalize_header(header))
                if not match:
                    continue
                year_value = int(match.group(1))
                candidates.append((year_value, idx2))
            if candidates:
                candidates.sort(reverse=True)
                _, chosen_idx = candidates[0]
                chosen_header = headers[chosen_idx]
                LOGGER.warning("JIF column %r not found; using %r", name, chosen_header)
                return chosen_idx, str(chosen_header)

        for alias in ("JIF 2024", "Journal Impact Factor"):
            alias_idx = _find_column(headers, alias)
            if alias_idx is not None:
                LOGGER.warning("JIF column %r not found; using %r", name, alias)
                return alias_idx, alias

    return None, name


def _normalize_header(value: str | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip().lower()
    return HEADER_SPACE_RE.sub(" ", text)


def _safe_cell(row: tuple[object, ...], idx: int | None) -> object | None:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _coerce_float(value: object | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_int(value: object | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _coerce_str(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_issn_candidates(candidates: Iterable[str | None]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for value in candidates:
        if value is None:
            continue
        key = normalize_issn(value)
        if not key or key in seen:
            continue
        normalized.append(key)
        seen.add(key)
    return normalized


def _update_index(
    index: Dict[str, Dict[str, object]],
    key: str,
    payload: Dict[str, object],
    ambiguous: set[str],
    label: str,
) -> bool:
    if key in ambiguous:
        return False
    existing = index.get(key)
    if existing:
        if not _payload_equivalent(existing, payload):
            LOGGER.warning("Duplicate JIF entry for %s %r; dropping ambiguous key", label, key)
            index.pop(key, None)
            ambiguous.add(key)
        return False
    index[key] = payload
    return True


def _payload_equivalent(left: Dict[str, object], right: Dict[str, object]) -> bool:
    keys = set(left.keys()) | set(right.keys())
    for key in keys:
        if key == "raw_title":
            continue
        if left.get(key) != right.get(key):
            return False
    return True


def _length_ratio_ok(a: str, b: str, max_delta: float) -> bool:
    if not a or not b:
        return False
    diff = abs(len(a) - len(b))
    denom = max(len(a), len(b))
    if denom == 0:
        return False
    return diff / denom <= max_delta
