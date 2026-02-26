"""Ranking dataset I/O (YAML config + CSV/TSV/XLSX loaders)."""

from __future__ import annotations

import csv
import os
import re
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, Mapping

import yaml  # type: ignore[import-untyped]
from openpyxl import load_workbook

from ai_bias_search.rankings.base import RankingConfig, RankingEntry, normalize_issn, normalize_title
from ai_bias_search.utils.logging import configure_logging

LOGGER = configure_logging()

HEADER_TOKENS = {
    "id",
    "title",
    "name",
    "acronym",
    "abbr",
    "shortname",
    "rank",
    "issn",
    "eissn",
    "issnl",
    "year",
    "edition",
}


def load_config(path: Path) -> RankingConfig:
    """Load a ranking YAML config from *path*."""

    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"Ranking config at {path} is not a mapping")
    cfg = RankingConfig.model_validate(raw)
    cfg.config_path = path
    if not cfg.dataset_path.is_absolute():
        cfg.dataset_path = (path.parent / cfg.dataset_path).resolve()
    return cfg


def _normalize_header_name(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def _resolve_column(
    headers: list[str] | None, spec: str | int | None, *, required: bool, label: str
) -> int | None:
    if spec is None:
        if required:
            raise ValueError(f"Missing required field mapping for {label}")
        return None
    if isinstance(spec, int):
        return spec
    if headers is None:
        raise ValueError(f"Column mapping for {label} requires headers (got {spec!r})")
    target = _normalize_header_name(str(spec))
    if not target:
        if required:
            raise ValueError(f"Empty column mapping for {label}")
        return None
    normalized = [_normalize_header_name(header) for header in headers]
    for idx, norm in enumerate(normalized):
        if norm == target:
            return idx
    if required:
        raise ValueError(f"Column {spec!r} for {label} not found in dataset headers")
    LOGGER.warning("Optional column %r for %s not found in dataset headers", spec, label)
    return None


def _iter_rows_csv(path: Path, *, delimiter: str, encoding: str) -> tuple[list[str] | None, Iterator[list[str]]]:
    handle = path.open("r", encoding=encoding, newline="")
    reader = csv.reader(handle, delimiter=delimiter)
    try:
        first = next(reader, None)
        if first is None:
            handle.close()
            return None, iter(())
        headers = [str(cell) if cell is not None else "" for cell in first]

        def gen() -> Iterator[list[str]]:
            try:
                for row in reader:
                    yield [str(cell) if cell is not None else "" for cell in row]
            finally:
                handle.close()

        return headers, gen()
    except Exception:
        handle.close()
        raise


def _is_header_row(row: list[str]) -> bool:
    hits = 0
    for cell in row:
        cell_value = str(cell).strip().lower()
        if not cell_value:
            continue
        tokens = set(re.sub(r"[^a-z]", " ", cell_value).split())
        if tokens & HEADER_TOKENS:
            hits += 1
    return hits >= 2


def _coerce_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _coerce_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _coerce_rank_value(value: object, cfg: RankingConfig) -> object | None:
    if value is None:
        return None
    if cfg.rank_value_type == "float":
        return _coerce_float(value)
    if cfg.rank_value_type == "int":
        return _coerce_int(value)
    if cfg.rank_value_type == "str":
        text = _coerce_str(value)
        if text is None:
            return None
        return text

    # auto
    if isinstance(value, (int, float)):
        return value
    text = _coerce_str(value)
    if text is None:
        return None
    as_float = _coerce_float(text)
    if as_float is not None and re.fullmatch(r"[-+]?\d+(\.\d+)?", text):
        return as_float
    return text


def _apply_rank_allowlist(value: object | None, cfg: RankingConfig) -> object | None:
    allowlist = cfg.rank_value_allowlist
    if not allowlist:
        return value
    if value is None:
        return None
    text = str(value).strip().upper().replace(" ", "")
    allowed = {item.strip().upper().replace(" ", "") for item in allowlist}
    return text if text in allowed else None


def _resolve_dataset_path(cfg: RankingConfig) -> Path:
    env_key = cfg.dataset_path_env
    if env_key:
        override = os.getenv(env_key)
        if override and override.strip():
            return Path(override.strip()).expanduser().resolve()
    return cfg.dataset_path


def resolve_dataset_path(cfg: RankingConfig) -> Path:
    """Resolve the dataset path for *cfg*, applying optional env var overrides."""

    return _resolve_dataset_path(cfg)


def load_dataset(cfg: RankingConfig) -> list[RankingEntry]:
    """Load a ranking dataset into `RankingEntry` records."""

    dataset_path = _resolve_dataset_path(cfg)
    if not dataset_path.exists():
        raise FileNotFoundError(f"Ranking dataset not found: {dataset_path}")

    if cfg.format in ("csv", "tsv"):
        delimiter = cfg.delimiter
        if delimiter is None:
            delimiter = "\t" if cfg.format == "tsv" else ","
        return _load_csv_like(dataset_path, cfg, delimiter=delimiter)
    if cfg.format == "xlsx":
        return _load_xlsx(dataset_path, cfg)

    raise ValueError(f"Unsupported ranking dataset format: {cfg.format}")


def _load_csv_like(path: Path, cfg: RankingConfig, *, delimiter: str) -> list[RankingEntry]:
    headers, rows_iter = _iter_rows_csv(path, delimiter=delimiter, encoding=cfg.encoding)
    has_header = cfg.has_header
    if has_header == "auto":
        has_header = bool(headers and _is_header_row(headers))

    if not has_header:
        # If the dataset has no header, treat the first row as data and require integer column specs.
        rows_iter = _prepend_first_row(headers, rows_iter)
        headers = None

    title_idx = _resolve_column(headers, cfg.fields.get("title"), required=True, label="title")
    rank_idx = _resolve_column(
        headers, cfg.fields.get("rank_value"), required=True, label="rank_value"
    )
    venue_key_idx = _resolve_column(
        headers, cfg.fields.get("venue_key"), required=False, label="venue_key"
    )
    issn_print_idx = _resolve_column(
        headers, cfg.fields.get("issn_print"), required=False, label="issn_print"
    )
    issn_online_idx = _resolve_column(
        headers, cfg.fields.get("issn_online"), required=False, label="issn_online"
    )
    issn_l_idx = _resolve_column(headers, cfg.fields.get("issn_l"), required=False, label="issn_l")
    rank_year_idx = _resolve_column(
        headers, cfg.fields.get("rank_year"), required=False, label="rank_year"
    )

    alias_indices = [
        idx
        for idx in (
            _resolve_column(headers, spec, required=False, label="title_alias") for spec in cfg.title_alias_fields
        )
        if idx is not None
    ]
    extra_indices: dict[str, int] = {}
    for key, spec in cfg.extra_fields.items():
        idx = _resolve_column(headers, spec, required=False, label=f"extra:{key}")
        if idx is not None:
            extra_indices[key] = idx

    entries: list[RankingEntry] = []
    for row in rows_iter:
        if not row:
            continue
        if not any(cell.strip() for cell in row):
            continue

        title = _safe_cell(row, title_idx)
        if not title:
            continue
        title_norm = normalize_title(title, cfg.normalization)
        if not title_norm:
            continue

        rank_value_raw = _safe_cell(row, rank_idx)
        rank_value = _apply_rank_allowlist(_coerce_rank_value(rank_value_raw, cfg), cfg)

        rank_year = _coerce_int(_safe_cell(row, rank_year_idx)) if rank_year_idx is not None else None
        if rank_year is None:
            rank_year = cfg.default_rank_year

        issn_print = (
            normalize_issn(
                _safe_cell(row, issn_print_idx),
                validate_checksum=cfg.validate_issn_checksum,
            )
            if issn_print_idx is not None
            else None
        )
        issn_online = (
            normalize_issn(
                _safe_cell(row, issn_online_idx),
                validate_checksum=cfg.validate_issn_checksum,
            )
            if issn_online_idx is not None
            else None
        )
        issn_l = (
            normalize_issn(
                _safe_cell(row, issn_l_idx),
                validate_checksum=cfg.validate_issn_checksum,
            )
            if issn_l_idx is not None
            else None
        )

        venue_key = _safe_cell(row, venue_key_idx) if venue_key_idx is not None else None
        if venue_key is None:
            venue_key = issn_l or issn_print or issn_online or title_norm

        extra: dict[str, Any] = {"edition": cfg.edition} if cfg.edition else {}
        title_aliases: list[str] = []
        for idx in alias_indices:
            alias = _safe_cell(row, idx)
            if alias:
                title_aliases.append(alias)
        if title_aliases:
            extra["_title_aliases"] = title_aliases
        for key, idx in extra_indices.items():
            extra[key] = _safe_cell(row, idx)

        entries.append(
            RankingEntry(
                venue_key=str(venue_key),
                title=title,
                title_norm=title_norm,
                issn_print=issn_print,
                issn_online=issn_online,
                issn_l=issn_l,
                rank_value=rank_value,
                rank_year=rank_year,
                source_id=cfg.id,
                extra=extra,
            )
        )
    LOGGER.info("Loaded ranking dataset %s (id=%s rows=%d)", path, cfg.id, len(entries))
    return entries


def _prepend_first_row(headers: list[str] | None, rows_iter: Iterator[list[str]]) -> Iterator[list[str]]:
    if headers is None:
        return rows_iter

    def gen() -> Iterator[list[str]]:
        yield headers
        yield from rows_iter

    return gen()


def _safe_cell(row: list[str], idx: int | None) -> str | None:
    if idx is None:
        return None
    if idx < 0 or idx >= len(row):
        return None
    text = str(row[idx]).strip()
    return text or None


def _load_xlsx(path: Path, cfg: RankingConfig) -> list[RankingEntry]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[cfg.sheet_name] if cfg.sheet_name and cfg.sheet_name in workbook.sheetnames else workbook.active

        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

        def col(spec: str | int | None, *, required: bool, label: str) -> int | None:
            return _resolve_column(headers, spec, required=required, label=label)

        title_idx = col(cfg.fields.get("title"), required=True, label="title")
        rank_idx = col(cfg.fields.get("rank_value"), required=True, label="rank_value")
        assert title_idx is not None
        assert rank_idx is not None
        venue_key_idx = col(cfg.fields.get("venue_key"), required=False, label="venue_key")
        issn_print_idx = col(cfg.fields.get("issn_print"), required=False, label="issn_print")
        issn_online_idx = col(cfg.fields.get("issn_online"), required=False, label="issn_online")
        issn_l_idx = col(cfg.fields.get("issn_l"), required=False, label="issn_l")
        rank_year_idx = col(cfg.fields.get("rank_year"), required=False, label="rank_year")

        alias_indices = [
            idx
            for idx in (col(spec, required=False, label="title_alias") for spec in cfg.title_alias_fields)
            if idx is not None
        ]
        extra_indices: dict[str, int] = {}
        for key, spec in cfg.extra_fields.items():
            idx = col(spec, required=False, label=f"extra:{key}")
            if idx is not None:
                extra_indices[key] = idx

        entries: list[RankingEntry] = []
        for row_values in rows:
            if not row_values:
                continue
            row = list(row_values)

            raw_title = row[title_idx] if title_idx is not None and title_idx < len(row) else None
            title = _coerce_str(raw_title)
            if not title:
                continue
            title_norm = normalize_title(title, cfg.normalization)
            if not title_norm:
                continue

            rank_value = _apply_rank_allowlist(_coerce_rank_value(row[rank_idx], cfg), cfg)
            rank_year = _coerce_int(row[rank_year_idx]) if rank_year_idx is not None else None
            if rank_year is None:
                rank_year = cfg.default_rank_year

            issn_print = (
                normalize_issn(_coerce_str(row[issn_print_idx]), validate_checksum=cfg.validate_issn_checksum)
                if issn_print_idx is not None
                else None
            )
            issn_online = (
                normalize_issn(_coerce_str(row[issn_online_idx]), validate_checksum=cfg.validate_issn_checksum)
                if issn_online_idx is not None
                else None
            )
            issn_l = (
                normalize_issn(_coerce_str(row[issn_l_idx]), validate_checksum=cfg.validate_issn_checksum)
                if issn_l_idx is not None
                else None
            )

            venue_key_raw = row[venue_key_idx] if venue_key_idx is not None and venue_key_idx < len(row) else None
            venue_key = _coerce_str(venue_key_raw) or issn_l or issn_print or issn_online or title_norm

            extra: dict[str, Any] = {"edition": cfg.edition} if cfg.edition else {}
            title_aliases: list[str] = []
            for idx in alias_indices:
                if idx >= len(row):
                    continue
                alias = _coerce_str(row[idx])
                if alias:
                    title_aliases.append(alias)
            if title_aliases:
                extra["_title_aliases"] = title_aliases
            for key, idx in extra_indices.items():
                if idx < len(row):
                    extra[key] = row[idx]

            entries.append(
                RankingEntry(
                    venue_key=str(venue_key),
                    title=title,
                    title_norm=title_norm,
                    issn_print=issn_print,
                    issn_online=issn_online,
                    issn_l=issn_l,
                    rank_value=rank_value,
                    rank_year=rank_year,
                    source_id=cfg.id,
                    extra=extra,
                )
            )

        LOGGER.info("Loaded ranking dataset %s (id=%s rows=%d)", path, cfg.id, len(entries))
        return entries
    finally:
        workbook.close()
