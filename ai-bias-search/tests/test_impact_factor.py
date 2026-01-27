from pathlib import Path

import pytest
from openpyxl import Workbook

from ai_bias_search.utils.config import ImpactFactorConfig
from ai_bias_search.utils.impact_factor import (
    ImpactFactorIndex,
    load_jif_xlsx,
    match_jcr_entry,
    normalize_issn,
)


def _write_jif_xlsx_full(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Journal",
            "JIF",
            "Year",
            "Publisher",
            "ISSN",
            "eISSN",
            "Total Cites",
            "Total Articles",
            "Citable Items",
            "5-Year JIF",
            "JIF Without Self-Cites",
            "JCI",
            "JIF Quartile",
            "JIF Rank",
        ]
    )
    ws.append(
        [
            "Journal of Testing",
            2.5,
            2024,
            "Test Pub",
            "1234-5678",
            "8765-4321",
            100,
            10,
            8,
            3.1,
            2.3,
            1.2,
            "Q1",
            "5/120",
        ]
    )
    ws.append(
        [
            "A & B Studies",
            3.0,
            2023,
            "Alpha Pub",
            "1111-2222",
            None,
            80,
            8,
            6,
            2.8,
            2.1,
            1.0,
            "Q2",
            "10/120",
        ]
    )
    ws.append(
        [
            "ABCD Journal",
            1.0,
            2022,
            "Delta Pub",
            "2222-3333",
            None,
            50,
            5,
            4,
            1.5,
            1.2,
            0.8,
            "Q3",
            "60/120",
        ]
    )
    ws.append(
        [
            "ABCE Journal",
            2.0,
            2022,
            "Echo Pub",
            "3333-4444",
            None,
            60,
            6,
            5,
            2.0,
            1.6,
            0.9,
            "Q3",
            "55/120",
        ]
    )
    wb.save(path)


def _write_jif_xlsx_minimal(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Journal", "JIF", "Year"])
    ws.append(["Journal of Testing", 2.5, 2024])
    wb.save(path)


def _write_jif_xlsx_jcr_export(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Rank",
            "Journal Name",
            "JCR Year",
            "Publisher",
            "ISSN",
            "eISSN",
            "JIF 2024",
            "5-Year JIF",
        ]
    )
    ws.append(
        [
            1,
            "Journal of Testing",
            2024,
            "Test Pub",
            "1234-5678",
            "8765-4321",
            2.5,
            3.1,
        ]
    )
    wb.save(path)


def _load_index(path: Path) -> ImpactFactorIndex:
    return load_jif_xlsx(
        path,
        sheet_name=None,
        title_column="Journal",
        jif_column="JIF",
        year_column="Year",
        publisher_column="Publisher",
        issn_column="ISSN",
        eissn_column="eISSN",
        total_cites_column="Total Cites",
        total_articles_column="Total Articles",
        citable_items_column="Citable Items",
        jif_5y_column="5-Year JIF",
        jif_wo_self_cites_column="JIF Without Self-Cites",
        jci_column="JCI",
        quartile_column="JIF Quartile",
        jif_rank_column="JIF Rank",
    )


def test_exact_match_from_normalized_title(tmp_path: Path) -> None:
    path = tmp_path / "jif.xlsx"
    _write_jif_xlsx_full(path)
    index = _load_index(path)
    cfg = ImpactFactorConfig(enabled=True, allow_fuzzy=False)
    payload = match_jcr_entry("A and B Studies", [], index, cfg)
    assert payload["impact_factor_match"] == "title_exact"
    assert payload["impact_factor"] == 3.0
    assert payload["impact_factor_year"] == 2023
    assert payload["jcr_publisher"] == "Alpha Pub"


def test_fuzzy_match_requires_flag_and_threshold(tmp_path: Path) -> None:
    path = tmp_path / "jif.xlsx"
    _write_jif_xlsx_full(path)
    index = _load_index(path)
    query = "Journal of Testin"

    cfg_disabled = ImpactFactorConfig(enabled=True, allow_fuzzy=False)
    payload = match_jcr_entry(query, [], index, cfg_disabled)
    assert payload["impact_factor"] is None
    assert payload["impact_factor_match"] == "none"

    cfg_enabled = ImpactFactorConfig(
        enabled=True,
        allow_fuzzy=True,
        fuzzy_threshold=85,
        min_title_len=1,
        max_len_ratio_delta=0.2,
    )
    payload = match_jcr_entry(query, [], index, cfg_enabled)
    assert payload["impact_factor_match"] == "title_fuzzy"
    assert payload["impact_factor"] == 2.5
    assert payload["impact_factor_year"] == 2024


def test_ambiguous_fuzzy_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "jif.xlsx"
    _write_jif_xlsx_full(path)
    index = _load_index(path)
    cfg = ImpactFactorConfig(
        enabled=True,
        allow_fuzzy=True,
        fuzzy_threshold=85,
        min_title_len=1,
        max_len_ratio_delta=0.2,
        reject_ambiguous=True,
    )
    payload = match_jcr_entry("ABCF Journal", [], index, cfg)
    assert payload["impact_factor_match"] == "ambiguous"
    assert payload["impact_factor"] is None
    assert payload["impact_factor_year"] is None


def test_issn_normalization() -> None:
    assert normalize_issn("12345678") == "1234-5678"
    assert normalize_issn("1234567") == "0123-4567"
    assert normalize_issn("1234 567x") == "1234-567X"
    assert normalize_issn("1234-567") is None
    assert normalize_issn("bad") is None


def test_issn_match_precedence_over_title(tmp_path: Path) -> None:
    path = tmp_path / "jif.xlsx"
    _write_jif_xlsx_full(path)
    index = _load_index(path)
    cfg = ImpactFactorConfig(
        enabled=True,
        allow_fuzzy=True,
        fuzzy_threshold=80,
        min_title_len=1,
        max_len_ratio_delta=0.2,
    )
    payload = match_jcr_entry("ABCD Journa", ["3333-4444"], index, cfg)
    assert payload["impact_factor_match"] == "issn_exact"
    assert payload["jcr_issn"] == "3333-4444"
    assert payload["impact_factor"] == 2.0


def test_eissn_index_match(tmp_path: Path) -> None:
    path = tmp_path / "jif.xlsx"
    _write_jif_xlsx_full(path)
    index = _load_index(path)
    cfg = ImpactFactorConfig(enabled=True, allow_fuzzy=False, min_title_len=1)
    payload = match_jcr_entry(None, ["8765-4321"], index, cfg)
    assert payload["impact_factor_match"] == "issn_exact"
    assert payload["jcr_eissn"] == "8765-4321"
    assert payload["impact_factor"] == 2.5


def test_missing_optional_columns_warn(tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
    path = tmp_path / "jif.xlsx"
    _write_jif_xlsx_minimal(path)
    caplog.set_level("WARNING")
    index = _load_index(path)
    cfg = ImpactFactorConfig(enabled=True, allow_fuzzy=False)
    payload = match_jcr_entry("Journal of Testing", [], index, cfg)
    assert payload["impact_factor"] == 2.5
    assert payload["jcr_publisher"] is None
    assert any("Publisher" in record.message for record in caplog.records)


def test_loader_falls_back_to_common_jcr_headers(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    path = tmp_path / "jif.xlsx"
    _write_jif_xlsx_jcr_export(path)
    caplog.set_level("WARNING")
    index = load_jif_xlsx(
        path,
        sheet_name=None,
        title_column="Journal",
        jif_column="JIF",
        year_column="Year",
        publisher_column="Publisher",
        issn_column="ISSN",
        eissn_column="eISSN",
        jif_5y_column="5-Year JIF",
    )
    cfg = ImpactFactorConfig(enabled=True, allow_fuzzy=False)
    payload = match_jcr_entry("Journal of Testing", [], index, cfg)
    assert payload["impact_factor_match"] == "title_exact"
    assert payload["impact_factor"] == 2.5
    assert payload["impact_factor_year"] == 2024
    assert any("JIF title column" in record.message for record in caplog.records)
    assert any("JIF column" in record.message for record in caplog.records)
    assert any("JIF year column" in record.message for record in caplog.records)
